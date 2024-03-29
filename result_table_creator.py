import pandas as pd
from openpyxl import Workbook
import openpyxl
import os
import getpass
import json
from copy import copy


jsonData = './build/json_data.json' 
resultPath = "C:\\Users\\"+getpass.getuser()+"\\Documents\\Расписание_Занятий\\"

# проверка наличия пути с результатами и его созадние, если нет
if(not os.path.exists(resultPath)):
    os.mkdir(resultPath)

def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)

def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print('Unable to copy default column wide')
    else:
        target_sheet.sheet_format.defaultColWidth = copy(source_sheet.sheet_format.defaultColWidth)

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(source_sheet.column_dimensions[key].min)   # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(source_sheet.column_dimensions[key].max)  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(source_sheet.column_dimensions[key].width) # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(source_sheet.column_dimensions[key].hidden)

def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)

def saveJsonToFile(timetables, file_name):
    if (file_name == "timetables.xlsx"):
        # загружаем template лист
        # template = parse_list(load_file('template.xls'), 0)

        # создаем новый Workbook
        wb_target = openpyxl.Workbook()

        #test = template.values[pointer_x][pointer_y]

        for timetable_name in timetables:
            target_sheet = wb_target.create_sheet(timetable_name)

            wb_source = openpyxl.load_workbook("template.xlsx", data_only=True)
            source_sheet = wb_source["template"]

            copy_sheet(source_sheet, target_sheet)

            # ws.column_dimensions['A'].width = 50
            # ws.column_dimensions['B'].width = 50
            # ws.column_dimensions['C'].width = 50
            # ws.column_dimensions['D'].width = 50
            # ws.column_dimensions['E'].width = 50

            timetable = timetables[timetable_name]
            day_shift = 3

            target_sheet.cell(row=day_shift,column=1).value = timetable_name
            # target_sheet.cell(row=day_shift+lesson_shift,column=2+1).value = i+1
            for idx, day in enumerate(timetable['odd']):
                lesson_shift = 0
                for i, lesson in enumerate(day):
                    if lesson == '':
                        continue
                    else:
                        # записываем номер пары
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+1).value = i+1
                        # записываем предмет
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+2).value = lesson["lesson"]
                        # записываем Тип зан.
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+4).value = 'лек'
                        # записываем преподавателя
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+5).value = lesson["teacher"]
                        # записываем аудиторию
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+6).value = lesson["cabinet"]
                        lesson_shift += 1
                day_shift += 6 # сдвигаемся на следующий день
            day_shift = 3
            for idx, day in enumerate(timetable['even']):
                lesson_shift = 0
                for i, lesson in enumerate(day):
                    if lesson == '':
                        continue
                    else:
                        # записываем номер пары
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+1).value = i+1
                        # записываем предмет
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+2).value = lesson["lesson"]
                        # записываем Тип зан.
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+4).value = 'лек'
                        # записываем преподавателя
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+5).value = lesson["teacher"]
                        # записываем аудиторию
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+6).value = lesson["cabinet"]
                        lesson_shift += 1
                day_shift += 6 # сдвигаемся на следующий день

        del wb_target['Sheet']
        wb_target.save(resultPath+file_name)
        wb_target.close()
    if (file_name == "teachers.xlsx"):
        wb_target = openpyxl.Workbook()
        for timetable_name in timetables:
            target_sheet = wb_target.create_sheet(timetable_name)
            wb_source = openpyxl.load_workbook("templateT.xlsx", data_only=True)
            source_sheet = wb_source["template"]
            copy_sheet(source_sheet, target_sheet)
            timetable = timetables[timetable_name]
            day_shift = 3
            target_sheet.cell(row=day_shift,column=1).value = timetable_name
            for day in timetable['odd']:
                lesson_shift = 0
                for i, lesson in enumerate(timetable['odd'][day]):
                    if not(day in timetables[timetable_name]['odd'] and lesson in timetables[timetable_name]['odd'][day]):
                        continue
                    else:
                        lesson_shift = int(lesson)
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+1).value = int(lesson)+1
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+2).value = timetables[timetable_name]['odd'][day][lesson]["lesson"]
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+4).value = 'лек'
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+5).value = timetables[timetable_name]['odd'][day][lesson]["group"]
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+6).value = timetables[timetable_name]['odd'][day][lesson]["cabinet"]
                day_shift += 6
            day_shift = 3
            for day in timetable['even']:
                lesson_shift = 0
                for i, lesson in enumerate(timetable['even'][day]):
                    if not(day in timetables[timetable_name]['even'] and lesson in timetables[timetable_name]['even'][day]):
                        continue
                    else:
                        lesson_shift = int(lesson)
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+1).value = int(lesson)+1
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+2).value = timetables[timetable_name]['even'][day][lesson]["lesson"]
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+4).value = 'лек'
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+5).value = timetables[timetable_name]['even'][day][lesson]["group"]
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+6).value = timetables[timetable_name]['even'][day][lesson]["cabinet"]
                day_shift += 6
        del wb_target['Sheet']
        wb_target.save(resultPath+file_name)
        wb_target.close()
    if (file_name == "audiences.xlsx"):
        wb_target = openpyxl.Workbook()
        for timetable_name in timetables:
            target_sheet = wb_target.create_sheet(timetable_name.replace('/', ' '))
            wb_source = openpyxl.load_workbook("templateA.xlsx", data_only=True)
            source_sheet = wb_source["template"]
            copy_sheet(source_sheet, target_sheet)
            timetable = timetables[timetable_name]
            day_shift = 3
            target_sheet.cell(row=day_shift,column=1).value = timetable_name
            for day in timetable['odd']:
                lesson_shift = 0
                for i, lesson in enumerate(timetable['odd'][day]):
                    if not(day in timetables[timetable_name]['odd'] and lesson in timetables[timetable_name]['odd'][day]):
                        continue
                    else:
                        lesson_shift = int(lesson)
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+1).value = int(lesson)+1
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+2).value = timetables[timetable_name]['odd'][day][lesson]["lesson"]
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+4).value = 'лек'
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+5).value = timetables[timetable_name]['odd'][day][lesson]["group"]
                        target_sheet.cell(row=day_shift+lesson_shift,column=2+6).value = timetables[timetable_name]['odd'][day][lesson]["cabinet"]
                day_shift += 6
            day_shift = 3
            for day in timetable['even']:
                lesson_shift = 0
                for i, lesson in enumerate(timetable['even'][day]):
                    if not(day in timetables[timetable_name]['even'] and lesson in timetables[timetable_name]['even'][day]):
                        continue
                    else:
                        lesson_shift = int(lesson)
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+1).value = int(lesson)+1
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+2).value = timetables[timetable_name]['even'][day][lesson]["lesson"]
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+4).value = 'лек'
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+5).value = timetables[timetable_name]['even'][day][lesson]["group"]
                        target_sheet.cell(row=day_shift+lesson_shift,column=9+6).value = timetables[timetable_name]['even'][day][lesson]["cabinet"]
                day_shift += 6
        del wb_target['Sheet']
        wb_target.save(resultPath+file_name)
        wb_target.close()
    
def loadTimetableFromJsonFile(filename):
    file = open(filename)
    data = file.read()
    file.close()
    result = json.loads(data)
    return result

def parse_list(file, list):
    try:
        list = int(list)
        # Загрузить лист по его имени
        return file.parse(file.sheet_names[list])
    except:
        print("Лист некорректен!")
        exit()

def load_file(in_file):
    # Загружаем spreadsheet в объект pandas
    file = pd.ExcelFile(in_file)
    return file

def main():
    data = loadTimetableFromJsonFile(jsonData)
    timetables = data['timetables']
    teachers_timetables = data['teachers_timetables']
    cabinets_timetables = data['cabinets_timetables']
    saveJsonToFile(cabinets_timetables, "audiences.xlsx")
    saveJsonToFile(teachers_timetables, "teachers.xlsx")
    saveJsonToFile(timetables, "timetables.xlsx")

if __name__ == '__main__':
    main()