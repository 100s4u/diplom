import pandas as pd
from openpyxl import Workbook
import sys

def getListsFromFile():
    # загружаем файл, переданный программе в ее аргументах
    file = load_file()

    # просим пользователя выбрать лист для работы программы
    lists = select_list(file)
    # на этом моменте получили целый лист, готовый к парсингу (selected_list)

    return lists
    
def saveTimetableToFile(timetables):
    wb = Workbook()
    ws = wb.create_sheet('Test')

    for timetable in timetables:
        timetable = timetables[timetable]
        ws.cell(row=1,column=5).value = 'Нечётная неделя' 
        for idx, day in enumerate(timetable['odd']):
            for i, lesson in enumerate(day):
                if lesson == '':
                    continue
                else:
                    ws.cell(row=idx+2,column=i+1).value = f'{i+1}. {lesson["lesson"]}'
        ws.cell(row=1,column=11).value = 'Чётная неделя'
        for idx, day in enumerate(timetable['even']):
            for i, lesson in enumerate(day):
                if lesson == '':
                    continue
                else:
                    ws.cell(row=idx+15,column=i+1).value = f'{i+1}. {lesson["lesson"]}'
    
    del wb['Sheet']
    wb.save("sample.xlsx")
    wb.close()

def select_list(file):
    # ввод листа от пользователя
    # selected_list = 0
    # if (len(xl.sheet_names) > 1):
    #     print("Обнаружено несколько листов. Выберите необходимый лист:")
    #     # Печатаем название листов в данном файле
    #     for idx, list_name in enumerate(xl.sheet_names):
    #         print(f"{idx+1}. {list_name}")
    #     selected_list = input("Введите число: ")

    # лист с парами
    lessons_list = 1 # второй по счету лист
    lessons_list = parse_list(file, lessons_list)

    # лист с заказными парами
    ordered_lessons_list = 2 # третий по счету лист
    ordered_lessons_list = parse_list(file, ordered_lessons_list)

    return [lessons_list, ordered_lessons_list]

def parse_list(file, selected_list):
    try:
        selected_list = int(selected_list)
        # Загрузить лист по его имени
        return file.parse(file.sheet_names[selected_list])
    except:
        print("Лист некорректен!")
        exit()

def load_file():
    # проверяем наличие переданного файла
    if (len(sys.argv) < 2):
        print("Ошибка: запуск без аргументов!")
        exit()
    else:
        in_file = sys.argv[1]

    # Загружаем spreadsheet в объект pandas
    file = pd.ExcelFile(in_file)
    return file